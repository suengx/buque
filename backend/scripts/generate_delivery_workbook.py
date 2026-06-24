"""生成一期交付验收确认表（与企业原始 workbook 风格对齐）。"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[2]
OUT_PATH = REPO_ROOT / "docs" / "delivery" / "补雀_一期交付验收确认表.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14)
SECTION_FONT = Font(bold=True, size=12)
WRAP = Alignment(wrap_text=True, vertical="top")


def style_header_row(ws, row: int, ncol: int) -> None:
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def set_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def sheet_usage(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "00_使用说明"
    ws["A1"] = "补雀 BuQue 一期交付验收确认表"
    ws["A1"].font = TITLE_FONT
    rows = [
        ("用途", "供计划/运营/仓储/管理层在试运行后，对照系统逐项确认是否达到一期交付标准。勾选 + 简短备注即可。"),
        ("填写人", "计划部主责；运营、仓储协同；IT/开发仅作演示支持，不参与业务结论勾选。"),
        (
            "建议节奏",
            "① 开发演示 30 分钟走查（见下方议程）→ ② 业务试用 3~5 个工作日 → ③ 本表集中评审签字。",
        ),
        (
            "与旧表关系",
            "本表继承原《05_一期范围_验收》结构，已按当前系统实现更新；指标定义保持通俗。",
        ),
        ("不算一期必须", "消息推送、运营计划导入、预测偏差规则、自动采购/调拨 — 已在范围表标注延后。"),
        ("演示会议程（约30分钟）", ""),
        ("1", "登录系统（邮箱或 Google）"),
        ("2", "日报总览：监控 SKU 数、红黄橙预警"),
        ("3", "风险预警：筛选 + 打开 SKU 详情（解释、DOS、建议）"),
        ("4", "监控助手：快捷提问「今天有哪些红色预警？」"),
        ("5", "采纳解释 + 人工反馈提交"),
        ("6", "顶栏手动「数据同步」触发日批"),
        ("7", "Q&A；现场填写 03_功能走查"),
    ]
    r = 3
    for k, v in rows:
        ws.cell(r, 1, k).font = Font(bold=True) if not str(k).isdigit() else None
        ws.cell(r, 2, v).alignment = WRAP
        r += 1
    set_widths(ws, [22, 88])


def sheet_scope(wb: Workbook) -> None:
    ws = wb.create_sheet("01_一期范围确认")
    ws["A1"] = "一、一期范围确认（实现对照）"
    ws["A1"].font = SECTION_FONT
    headers = ["范围项", "是否纳入一期", "系统对应功能（通俗）", "业务确认", "确认人", "确认日期", "备注"]
    ws.append(headers)
    style_header_row(ws, 2, len(headers))
    for row in [
        ("每日销量与库存自动监控", "是", "顶栏「数据同步」；每日 06:00 可自动跑（生产环境）", "", "", "", "核心"),
        ("断货/滞销/销量异常预警", "是", "「风险预警」清单 + 日报红黄橙统计", "", "", "", "核心"),
        ("规则解释（非对话）", "是", "清单/详情「主解释」— 规则表自动生成，非 Chat", "", "", "", ""),
        ("日报总览", "是", "首页「日报总览」", "", "", "", ""),
        ("单 SKU 分析卡", "是", "预警清单 → 查看详情", "", "", "", ""),
        ("监控助手对话", "是", "「监控助手」追问、采纳解释", "", "", "", "唯一 LLM 入口"),
        ("人工反馈", "是", "「人工反馈」页", "", "", "", ""),
        ("规则参数可调", "是", "「规则配置」页", "", "", "", ""),
        ("红/橙消息推送", "否（延后）", "未实现企微/邮件", "□同意延后", "", "", ""),
        ("运营计划导入", "否（延后）", "未实现", "□同意延后", "", "", ""),
        ("预测偏差预警", "否（延后）", "规则默认关闭", "□同意延后", "", "", ""),
        ("自动改预测/采购/调拨", "否", "只读建议，不写回 ERP", "□确认边界", "", "", ""),
    ]:
        ws.append(list(row))
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=4, max_col=4):
        if not row[0].value or row[0].value == "":
            row[0].value = "□通过  □不通过  □待确认"
    set_widths(ws, [24, 12, 38, 22, 10, 12, 14])


def sheet_metrics(wb: Workbook) -> None:
    ws = wb.create_sheet("02_验收指标")
    ws["A1"] = "二、一期验收指标（请业务填写结论）"
    ws["A1"].font = SECTION_FONT
    headers = ["指标类别", "指标名称（通俗）", "建议目标", "实际结论", "是否达标", "统计方式", "填写人", "备注"]
    ws.append(headers)
    style_header_row(ws, 2, len(headers))
    for m in [
        ("规则一致性", "与 Excel 监控结果一致比例", "≥95%", "", "□是 □否 □暂无法测", "compare_excel.py + 业务基准表", "", ""),
        ("预警有用性", "红/橙灯值得处理的比例", "建议≥70%", "", "□是 □否 □试用中", "试用人工复核", "", ""),
        ("误报控制", "标异常但不用管的比例", "逐步下降", "", "□可接受 □偏高", "试用反馈", "", ""),
        ("日报时效", "上班前能看到当日日报", "按约定", "", "□是 □否", "快照完成时间/日志", "", "生产 06:00"),
        ("分析提效", "盯表初筛时间减少", "建议降50%+", "", "□是 □否 □待观察", "访谈/工时", "", "建议项"),
        ("建议采纳", "建议被采纳比例", "建议≥50%", "", "□是 □否 □样本少", "反馈页记录", "", "建议项"),
        ("系统稳定", "同步分析能跑完", "≥99%", "", "□是 □否", "erp_sync_job + 日志", "", ""),
    ]:
        ws.append(list(m))
    set_widths(ws, [14, 26, 12, 14, 18, 28, 10, 14])


def sheet_uat(wb: Workbook) -> None:
    ws = wb.create_sheet("03_功能走查")
    ws["A1"] = "三、功能走查（演示/试用时逐项打勾）"
    ws["A1"].font = SECTION_FONT
    headers = ["序号", "验证事项（通俗）", "操作路径", "结果", "问题记录"]
    ws.append(headers)
    style_header_row(ws, 2, len(headers))
    for u in [
        ("1", "生产/试运行 URL 可访问（HTTPS）", "浏览器打开约定域名", "□通过 □失败", ""),
        ("2", "能看到监控 SKU 数与预警总量", "日报总览", "□通过 □失败", ""),
        ("3", "能按仓库/等级/类型筛预警", "风险预警 → 筛选", "□通过 □失败", ""),
        ("4", "SKU 详情有 DOS、解释、建议", "查看详情", "□通过 □失败", ""),
        ("5", "主解释可读（非乱码）", "详情主解释区", "□通过 □失败", "规则解释非 Chat"),
        ("6", "能问「今天红色预警有哪些」", "监控助手", "□通过 □失败", ""),
        ("7", "能问某 SKU 库存销量", "监控助手", "□通过 □失败", ""),
        ("8", "解释草稿可采纳", "对话内采纳", "□通过 □失败", ""),
        ("9", "能提交人工反馈", "人工反馈", "□通过 □失败", ""),
        ("10", "能切换历史快照", "顶栏数据快照", "□通过 □失败", ""),
        ("11", "能手动触发同步+分析", "顶栏数据同步", "□通过 □失败", ""),
        ("12", "换快照后会话列表联动", "监控助手", "□通过 □失败", ""),
        ("13", "每日定时任务有执行记录", "次日查日报/日志", "□通过 □失败", "生产 06:00"),
    ]:
        ws.append(list(u))
    set_widths(ws, [6, 34, 24, 16, 30])


def sheet_excel(wb: Workbook) -> None:
    ws = wb.create_sheet("04_Excel一致率")
    ws["A1"] = "四、Excel 规则一致率对照（M2 硬指标）"
    ws["A1"].font = SECTION_FONT
    ws["A3"] = "说明：计划部提供 excel_baseline.xlsx 后，IT 执行："
    ws["A4"] = "uv run python scripts/compare_excel.py --baseline fixtures/excel_baseline.xlsx --system <系统导出> --keys sku,warehouse"
    headers = ["SKU", "仓库", "Excel风险等级", "系统风险等级", "是否一致", "备注"]
    ws.append([])
    ws.append(headers)
    style_header_row(ws, 6, len(headers))
    for _ in range(20):
        ws.append(["", "", "", "", "□是 □否", ""])
    ws["A28"] = "汇总一致率"
    ws["B28"] = ""
    ws["C28"] = "目标 ≥95%"
    set_widths(ws, [16, 18, 14, 14, 12, 24])


def sheet_trial(wb: Workbook) -> None:
    ws = wb.create_sheet("05_试运行记录")
    ws["A1"] = "五、试运行周报（每周或每日填写）"
    ws["A1"].font = SECTION_FONT
    headers = ["日期", "同步是否成功", "快照ID", "红/橙/黄数量", "主要问题", "处理人", "状态"]
    ws.append(headers)
    style_header_row(ws, 2, len(headers))
    for _ in range(14):
        ws.append(["", "□成功 □失败", "", "", "", "", "□已解决 □进行中"])
    set_widths(ws, [12, 16, 10, 14, 32, 10, 16])


def sheet_sign(wb: Workbook) -> None:
    ws = wb.create_sheet("06_交付签字")
    ws["A1"] = "六、一期交付结论"
    ws["A1"].font = SECTION_FONT
    ws["A3"] = "综合结论（勾选一项）"
    ws["A4"] = "□ 同意一期交付验收通过，进入试运行扩大范围"
    ws["A5"] = "□ 原则通过，遗留项见各表备注，不影响试运行"
    ws["A6"] = "□ 暂不通过，需整改后复验"
    ws["A8"] = "遗留项摘要"
    ws["A10"] = "角色"
    ws["B10"] = "姓名"
    ws["C10"] = "签字"
    ws["D10"] = "日期"
    for c in range(1, 5):
        ws.cell(10, c).font = Font(bold=True)
    for i, role in enumerate(["计划部", "运营部", "仓储/物流", "IT/开发", "项目负责人"], 11):
        ws.cell(i, 1, role)
    set_widths(ws, [16, 16, 16, 14])


def sheet_faq(wb: Workbook) -> None:
    ws = wb.create_sheet("07_补充说明_白话")
    ws["A1"] = "七、常见疑问（给业务）"
    ws["A1"].font = SECTION_FONT
    faqs = [
        ("多粒度监控？", "后台按全公司/单仓库/单平台分别计算；一期页面默认只看「单仓库」清单，与 Excel 盯仓习惯一致。"),
        ("规则解释是不是 AI？", "日报和清单上的主解释是规则程序生成的，不是 Chat。只有监控助手里主动提问才用 AI。"),
        ("定时任务什么时候跑？", "生产环境每天 06:00（上海时区）自动从 ERP 拉数并分析；也可顶栏手动同步。"),
        ("和 Excel 怎么验收？", "提供 Excel 基准表，IT 跑对照脚本，一致率填入 02、04 表。"),
    ]
    r = 3
    for q, a in faqs:
        ws.cell(r, 1, q).font = Font(bold=True)
        ws.cell(r, 1).alignment = WRAP
        ws.cell(r, 2, a).alignment = WRAP
        r += 2
    set_widths(ws, [22, 80])


def main() -> None:
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    sheet_usage(wb)
    sheet_scope(wb)
    sheet_metrics(wb)
    sheet_uat(wb)
    sheet_excel(wb)
    sheet_trial(wb)
    sheet_sign(wb)
    sheet_faq(wb)
    wb.save(OUT_PATH)
    print(f"已生成: {OUT_PATH}")


if __name__ == "__main__":
    main()
